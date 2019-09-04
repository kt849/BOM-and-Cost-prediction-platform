#!/usr/bin/env python
# coding: utf-8

# In[1]:


import numpy as np
import pandas as pd
from sklearn import preprocessing
from datascience import *
import matplotlib
import sys
from pyexcel.cookbook import merge_all_to_a_book
import glob
import openpyxl
import json

# In[30]:


Graphics_table= Table.read_table("Graphics_final.csv")


# In[31]:


Screen_table= Table.read_table("Screen_final.csv")


# In[32]:


RAM_table= Table.read_table("RAM_final.csv")


# In[34]:


SSD_table= Table.read_table("SSD_final.csv")


# In[35]:


Proc_table= Table.read_table("Processor_final.csv")


# In[36]:


Supplier_rating= Table.read_table("HDD_finall.csv")


# In[37]:


#converting into standard units
def standard_units(xyz):
    return(xyz-np.mean(xyz))/np.std(xyz)


# In[38]:


Graphic_rating= Graphics_table.with_columns('Normalization Time', standard_units(Graphics_table.column('Time')), 'Normalization Finance', standard_units(Graphics_table.column('Finance')), 'Normalization Quality', standard_units(Graphics_table.column('Quality')))


# In[39]:


Screen_rating= Screen_table.with_columns('Normalization Time', standard_units(Screen_table.column('Time')), 'Normalization Finance', standard_units(Screen_table.column('Finance')), 'Normalization Quality', standard_units(Screen_table.column('Quality')))


# In[40]:


RAM_rating= RAM_table.with_columns('Normalization Time', standard_units(RAM_table.column('Time')), 'Normalization Finance', standard_units(RAM_table.column('Finance')), 'Normalization Quality', standard_units(RAM_table.column('Quality')))


# In[41]:


Proc_rating= Proc_table.with_columns('Normalization Time', standard_units(RAM_table.column('Time')), 'Normalization Finance', standard_units(RAM_table.column('Finance')), 'Normalization Quality', standard_units(RAM_table.column('Quality')))


# In[42]:


SSD_rating= SSD_table.with_columns('Normalization Time', standard_units(SSD_table.column('Time')), 'Normalization Finance', standard_units(SSD_table.column('Finance')), 'Normalization Quality', standard_units(SSD_table.column('Quality')))


# In[43]:


Supplier_rating= Supplier_rating.with_columns('Normalization Time', standard_units(Supplier_rating.column('Time')), 'Normalization Finance', standard_units(Supplier_rating.column('Finance')), 'Normalization Quality', standard_units(Supplier_rating.column('Quality')))


# In[44]:


Graphic_rating


# In[45]:


RAM_rating


# In[46]:


SSD_rating


# In[47]:


Proc_rating


# In[48]:


Screen_rating


# In[49]:


Supplier_rating


# In[50]:


Graphic_x= Graphic_rating.column('Normalization Time')+Graphic_rating.column('Normalization Finance')+Graphic_rating.column('Normalization Quality')


# In[51]:


Screen_x= Screen_rating.column('Normalization Time')+Screen_rating.column('Normalization Finance')+Screen_rating.column('Normalization Quality')


# In[52]:


Proc_x= Proc_rating.column('Normalization Time')+Proc_rating.column('Normalization Finance')+Proc_rating.column('Normalization Quality')


# In[53]:


SSD_x= SSD_rating.column('Normalization Time')+SSD_rating.column('Normalization Finance')+SSD_rating.column('Normalization Quality')


# In[54]:


RAM_x= RAM_rating.column('Normalization Time')+RAM_rating.column('Normalization Finance')+RAM_rating.column('Normalization Quality')


# In[55]:


HDD_x= Supplier_rating.column('Normalization Time')+Supplier_rating.column('Normalization Finance')+Supplier_rating.column('Normalization Quality')


# In[56]:


Graphic_rating= Graphic_rating.with_column('x_val', Graphic_x)


# In[57]:


RAM_rating= RAM_rating.with_column('x_val', RAM_x)


# In[58]:


Proc_rating= Proc_rating.with_column('x_val', RAM_x)


# In[59]:


SSD_rating= SSD_rating.with_column('x_val', SSD_x)


# In[60]:


Screen_rating= Screen_rating.with_column('x_val', Screen_x)


# In[61]:


Supplier_rating= Supplier_rating.with_column('x_val', HDD_x)


# In[62]:


Graphic_rating


# In[63]:


RAM_rating


# In[64]:


Proc_rating


# In[65]:


Screen_rating


# In[66]:


SSD_rating


# In[67]:


Supplier_rating


# In[68]:


def predict_gy(x_val):
    nearby= Graphic_rating.where('x_val', are.between(x_val-1, x_val+1))
    return np.mean((nearby.column('Time')+nearby.column('Finance')+nearby.column('Quality'))/3)
    


# In[69]:


def predict_ry(x_val):
    nearby= RAM_rating.where('x_val', are.between(x_val-1, x_val+1))
    return np.mean((nearby.column('Time')+nearby.column('Finance')+nearby.column('Quality'))/3)
    


# In[70]:


def predict_py(x_val):
    nearby= Proc_rating.where('x_val', are.between(x_val-1, x_val+1))
    return np.mean((nearby.column('Time')+nearby.column('Finance')+nearby.column('Quality'))/3)


# In[71]:


def predict_sdy(x_val):
    nearby= SSD_rating.where('x_val', are.between(x_val-1, x_val+1))
    return np.mean((nearby.column('Time')+nearby.column('Finance')+nearby.column('Quality'))/3)


# In[72]:


def predict_sy(x_val):
    nearby= Screen_rating.where('x_val', are.between(x_val-1, x_val+1))
    return np.mean((nearby.column('Time')+nearby.column('Finance')+nearby.column('Quality'))/3)


# In[73]:


#predicting value of Rating given a particular x_value
def predict_y(x_val):
    nearby_points = Supplier_rating.where('x_val', are.between(x_val-1, x_val+1))
    return np.mean((nearby_points.column('Time')+nearby_points.column('Finance')+nearby_points.column('Quality'))/3)


# In[74]:


Graphic_rating= Graphic_rating.with_column('Predicted Rating', Graphic_rating.apply(predict_gy, 'x_val'))


# In[75]:


RAM_rating= RAM_rating.with_column('Predicted Rating', RAM_rating.apply(predict_ry, 'x_val'))


# In[76]:


SSD_rating= SSD_rating.with_column('Predicted Rating', SSD_rating.apply(predict_sdy, 'x_val'))


# In[77]:


Proc_rating= Proc_rating.with_column('Predicted Rating', SSD_rating.apply(predict_sdy, 'x_val'))


# In[78]:


Screen_rating= Screen_rating.with_column('Predicted Rating', Screen_rating.apply(predict_sy, 'x_val'))


# In[79]:


Supplier_rating= Supplier_rating.with_column('Predicted Rating', Supplier_rating.apply(predict_y, 'x_val'))


# In[80]:


Graphic_rating.sort("Predicted Rating", descending="True")


# In[81]:


RAM_rating.sort("Predicted Rating", descending="True")


# In[82]:


Proc_rating.sort("Predicted Rating", descending="True")


# In[83]:


SSD_rating.sort("Predicted Rating", descending="True")


# In[84]:


Screen_rating.sort("Predicted Rating", descending="True")


# In[85]:


Supplier_rating.sort("Predicted Rating", descending="True")


# In[87]:


def value_Graphic(n,x):
    
    tab= Graphic_rating.where("size", are.equal_to(n)).where("type", are.equal_to(x))
    d=tab.column("budget").take(1)
    tab1= tab.where("price", are.below(d))
    
    
    
    tab2= tab1.sort("Predicted Rating", descending="True")
    anu = tab2.drop('size','quality','Time','Quality','Finance','Normalization Time','Normalization Finance','Normalization Quality','x_val')
    anu = anu.sort('Predicted Rating',descending=True)
    if(len(anu.column("Predicted Rating"))>5):
        anu=anu.take(np.arange(0,5)).sort("price")
    elif(len(anu.column("Predicted Rating"))<=5):
        anu= anu.take(np.arange(0, len(anu.column("Predicted Rating")))).sort("price")
    else:
        print("No Value Found")
    anu= anu.select("type","Company","price","Predicted Rating")
    a= anu.to_csv("dell1.csv")
    return anu

value_Graphic(int(sys.argv[1]),int(sys.argv[2]))


# In[89]:


def value_Proc(n,x):
    tab= Proc_rating.where("size", are.equal_to(n)).where("type", are.equal_to(x))
    d=tab.column("budget").take(1)
    tab1= tab.where("price", are.below(d))
    
    tab2= tab1.sort("Predicted Rating", descending="True")
    anu = tab2.drop('size','quality','Time','Quality','Finance','Normalization Time','Normalization Finance','Normalization Quality','x_val')
    anu = anu.sort('Predicted Rating',descending=True)
    if(len(anu.column("Predicted Rating"))>5):
        anu=anu.take(np.arange(0,5))
    elif(len(anu.column("Predicted Rating"))<=5):
        anu= anu.take(np.arange(0, len(anu.column("Predicted Rating"))))
    else:
        print("No Value Found")
    anu= anu.select("type","Company","price","Predicted Rating")
    a= anu.to_csv("dell2.csv")
    return anu

value_Proc(int(sys.argv[3]),int(sys.argv[4]))


# In[90]:


def value_SSD(n,x):
   
    tab= SSD_rating.where("size", are.equal_to(n)).where("type", are.equal_to(x))
    d=tab.column("budget").take(1)
    tab1= tab.where("price", are.below(d))
    
    tab2= tab1.sort("Predicted Rating", descending="True")
    anu = tab2.drop('size','quality','Time','Quality','Finance','Normalization Time','Normalization Finance','Normalization Quality','x_val')
    anu = anu.sort('Predicted Rating',descending=True)

    if(len(anu.column("Predicted Rating"))>5):
        anu=anu.take(np.arange(0,5))
    elif(len(anu.column("Predicted Rating"))<=5):
        anu= anu.take(np.arange(0, len(anu.column("Predicted Rating"))))
    else:
        print("No Value Found")
    anu= anu.select("type","Company","price","Predicted Rating")
    a= anu.to_csv("dell3.csv")
    return anu

value_SSD(int(sys.argv[5]),int(sys.argv[6]))


# In[94]:


def value_RAM(n,x):
    tab= RAM_rating.where("size", are.equal_to(n)).where("type", are.equal_to(x))
    d=tab.column("budget").take(1)
    tab1= tab.where("price", are.below(d))
    
    tab2= tab1.sort("Predicted Rating", descending="True")
    anu = tab2.drop('size','quality','Time','Quality','Finance','Normalization Time','Normalization Finance','Normalization Quality','x_val')
    anu = anu.sort('Predicted Rating',descending=True)
    if(len(anu.column("Predicted Rating"))>5):
        anu=anu.take(np.arange(0,5))
    elif(len(anu.column("Predicted Rating"))<=5):
        anu= anu.take(np.arange(0, len(anu.column("Predicted Rating"))))
    else:
        print("No Value Found")
    anu= anu.select("type","Company","price","Predicted Rating")
    a= anu.to_csv("dell4.csv")
    return anu

value_RAM(int(sys.argv[7]),int(sys.argv[8]))


# In[92]:


def value_Screen(n,x):
    
    tab= Screen_rating.where("size", are.equal_to(n)).where("type", are.equal_to(x))
    d=tab.column("budget").take(1)
    tab1= tab.where("price", are.below(d))
    
    tab2= tab1.sort("Predicted Rating", descending="True")
    anu = tab2.drop('size','quality','Time','Quality','Finance','Normalization Time','Normalization Finance','Normalization Quality','x_val')
    anu = anu.sort('Predicted Rating',descending=True)
    if(len(anu.column("Predicted Rating"))>5):
        anu=anu.take(np.arange(0,5))
    elif(len(anu.column("Predicted Rating"))<=5):
        anu= anu.take(np.arange(0, len(anu.column("Predicted Rating"))))
    else:
        print("No Value Found")
    anu= anu.select("type","Company","price","Predicted Rating")
    a= anu.to_csv("dell5.csv")
    return anu

value_Screen(int(sys.argv[9]),int(sys.argv[10]))


# In[93]:


def value_HDD(n,x):
    tab= Supplier_rating.where("size", are.equal_to(n)).where("type", are.equal_to(x))
    d=tab.column("budget").take(1)
    tab1= tab.where("price", are.below(d))
    
    tab2= tab1.sort("Predicted Rating", descending="True")
    anu = tab2.drop('size','quality','Time','Quality','Finance','Normalization Time','Normalization Finance','Normalization Quality','x_val')
    anu = anu.sort('Predicted Rating',descending=True)
    if(len(anu.column("Predicted Rating"))>5):
        anu=anu.take(np.arange(0,5))
    elif(len(anu.column("Predicted Rating"))<=5):
        anu= anu.take(np.arange(0, len(anu.column("Predicted Rating"))))
    else:
        print("No Value Found")
    anu= anu.select("type","Company","price","Predicted Rating")
    a= anu.to_csv("dell6.csv")

    return anu


value_HDD(int(sys.argv[11]),int(sys.argv[12]))


merge_all_to_a_book(glob.glob("./dell1.csv"), "dell1.xlsx")
merge_all_to_a_book(glob.glob("./dell2.csv"), "dell2.xlsx")
merge_all_to_a_book(glob.glob("./dell3.csv"), "dell3.xlsx")
merge_all_to_a_book(glob.glob("./dell4.csv"), "dell4.xlsx")
merge_all_to_a_book(glob.glob("./dell5.csv"), "dell5.xlsx")
merge_all_to_a_book(glob.glob("./dell6.csv"), "dell6.xlsx")


config = {}
wb = openpyxl.load_workbook('./dell1.xlsx')
ws = wb.active

graphics = []
tmp = 2
while(tmp<7):
    arr = []
    arr.append(ws['A'+str(tmp)].value)
    arr.append(ws['B'+str(tmp)].value)
    arr.append(ws['C'+str(tmp)].value)
    arr.append(ws['D'+str(tmp)].value)
    tmp = tmp+1
    graphics.append(arr)


wb = openpyxl.load_workbook('./dell2.xlsx')
ws = wb.active

cpu = []
tmp = 2
while(tmp<7):
    arr = []
    arr.append(ws['A'+str(tmp)].value)
    arr.append(ws['B'+str(tmp)].value)
    arr.append(ws['C'+str(tmp)].value)
    arr.append(ws['D'+str(tmp)].value)
    tmp = tmp+1
    cpu.append(arr)


wb = openpyxl.load_workbook('./dell3.xlsx')
ws = wb.active

ssd = []
tmp = 2
while(tmp<7):
    arr = []
    arr.append(ws['A'+str(tmp)].value)
    arr.append(ws['B'+str(tmp)].value)
    arr.append(ws['C'+str(tmp)].value)
    arr.append(ws['D'+str(tmp)].value)
    tmp = tmp+1
    ssd.append(arr)


wb = openpyxl.load_workbook('./dell4.xlsx')
ws = wb.active

ram = []
tmp = 2
while(tmp<7):
    arr = []
    arr.append(ws['A'+str(tmp)].value)
    arr.append(ws['B'+str(tmp)].value)
    arr.append(ws['C'+str(tmp)].value)
    arr.append(ws['D'+str(tmp)].value)
    tmp = tmp+1
    ram.append(arr)


wb = openpyxl.load_workbook('./dell5.xlsx')
ws = wb.active

screen = []
tmp = 2
while(tmp<7):
    arr = []
    arr.append(ws['A'+str(tmp)].value)
    arr.append(ws['B'+str(tmp)].value)
    arr.append(ws['C'+str(tmp)].value)
    arr.append(ws['D'+str(tmp)].value)
    tmp = tmp+1
    screen.append(arr)


wb = openpyxl.load_workbook('./dell6.xlsx')
ws = wb.active

hdd = []
tmp = 2
while(tmp<7):
    arr = []
    arr.append(ws['A'+str(tmp)].value)
    arr.append(ws['B'+str(tmp)].value)
    arr.append(ws['C'+str(tmp)].value)
    arr.append(ws['D'+str(tmp)].value)
    tmp = tmp+1
    hdd.append(arr)


config["graphics"] = graphics
config["cpu"] = cpu
config["ssd"] = ssd
config["ram"] = ram
config["screen"] = screen
config["hdd"] = hdd




with open('opti.json', 'w') as json_file:
    json.dump(config, json_file)

print("acdsfdwgvdvsexybitchssf")
# print(config)
# In[ ]:




