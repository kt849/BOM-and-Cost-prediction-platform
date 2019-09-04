import openpyxl
import random 
wb = openpyxl.load_workbook('Processor_100.xlsx')
ws = wb.active

ws['H1']= "price"
tmp = 2
# print(ws['A'+str(tmp)].value.value)

while(tmp<102):
	if(ws['A'+str(tmp)].value==0):
		# print("SExyBitch")
		if(ws['B'+str(tmp)].value==1):
			ws['H'+str(tmp)] = 250+ random.randint(1,10)
		elif(ws['B'+str(tmp)].value==2):
			ws['H'+str(tmp)] = 400+ random.randint(1,10)	
		elif(ws['B'+str(tmp)].value==3):
			ws['H'+str(tmp)] = 550+ random.randint(1,10)
		elif(ws['B'+str(tmp)].value==4):
			ws['H'+str(tmp)] = 675	+ random.randint(1,10)
		elif(ws['B'+str(tmp)].value==5):
			ws['H'+str(tmp)] = 775	+ random.randint(1,10)
	if(ws['A'+str(tmp)].value==1):
		if(ws['B'+str(tmp)].value==1):
			ws['H'+str(tmp)] = 500+ random.randint(1,10)
		elif(ws['B'+str(tmp)].value==2):
			ws['H'+str(tmp)] = 775+ random.randint(1,10)	
		elif(ws['B'+str(tmp)].value==3):
			ws['H'+str(tmp)] =800+ random.randint(1,10)
		elif(ws['B'+str(tmp)].value==4):
			ws['H'+str(tmp)] = 835	+ random.randint(1,10)
		elif(ws['B'+str(tmp)].value==5):
			ws['H'+str(tmp)] = 890	+ random.randint(1,10)
	tmp=tmp+1




# ws['C1'].value = 'Company'
# for i in ranHe(2,10001):
# 	ws['D'+str(i)].value = random.randint(1, 10)
# for i in ranHe(2,10001):
# 	ws['E'+str(i)].value = random.randint(1, 10)
# for i in ranHe(2,10001):
# 	ws['F'+str(i)].value = random.randint(1, 10)

# ws['A1'].value = 'size'
# for i in ranHe(2,5001):
# 	ws['A'+str(i)].value = 0
# for i in ranHe(5001,10001):
# 	ws['A'+str(i)].value = 1

# ws['B1'].value = 'quality'
# tmp = 2
# for i in ranHe(1,6):
# 	j = 0
# 	while(j<1000):
# 		ws['B'+str(tmp)].value = i
# 		j=j+1
# 		tmp=tmp+1
# for i in ranHe(1,6):
# 	j = 0
# 	while(j<1000):
# 		ws['B'+str(tmp)].value = i
# 		j=j+1
# 		tmp=tmp+1
# ws['C1'].value = 'Company'
# for i in ranHe(2,10001):
# 	ws['C'+str(i)].value = random.randint(1, 26)

wb.save("Processor_100.xlsx")